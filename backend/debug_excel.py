"""Debug script to test PDF to Excel with image embedding."""
import fitz
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import os

def test_pdf_to_excel_with_images(pdf_path, output_path):
    print(f"Opening PDF: {pdf_path}")
    doc = fitz.open(pdf_path)
    wb = Workbook()
    
    for page_num, page in enumerate(doc):
        print(f"\n=== Page {page_num + 1} ===")
        
        if page_num == 0:
            ws = wb.active
            ws.title = "Page 1"
        else:
            ws = wb.create_sheet(title=f"Page {page_num + 1}")
        
        current_row = 1
        
        # Extract text
        text_dict = page.get_text("dict")
        text_count = 0
        for block in text_dict["blocks"]:
            if block["type"] == 0:
                for line in block["lines"]:
                    line_text = "".join(span["text"] for span in line["spans"])
                    if line_text.strip():
                        ws.cell(row=current_row, column=1, value=line_text.strip())
                        current_row += 1
                        text_count += 1
        
        print(f"  Text lines extracted: {text_count}")
        
        # Extract images
        current_row += 1
        image_list = page.get_images(full=True)
        print(f"  Images found on page: {len(image_list)}")
        
        for img_index, img_info in enumerate(image_list):
            try:
                xref = img_info[0]
                print(f"    Image {img_index + 1}: xref={xref}")
                
                base_image = doc.extract_image(xref)
                image_bytes = base_image["image"]
                image_ext = base_image["ext"]
                print(f"      Format: {image_ext}, Size: {len(image_bytes)} bytes")
                
                # Create image object
                img_stream = BytesIO(image_bytes)
                xl_img = XLImage(img_stream)
                print(f"      Dimensions: {xl_img.width}x{xl_img.height}")
                
                # Scale if needed
                max_width = 500
                max_height = 400
                width_ratio = max_width / xl_img.width if xl_img.width > max_width else 1
                height_ratio = max_height / xl_img.height if xl_img.height > max_height else 1
                scale = min(width_ratio, height_ratio)
                
                if scale < 1:
                    xl_img.width = int(xl_img.width * scale)
                    xl_img.height = int(xl_img.height * scale)
                    print(f"      Scaled to: {xl_img.width}x{xl_img.height}")
                
                cell_ref = f"A{current_row}"
                ws.add_image(xl_img, cell_ref)
                print(f"      Added at cell: {cell_ref}")
                
                rows_needed = max(1, int(xl_img.height / 15) + 1)
                current_row += rows_needed
                
            except Exception as e:
                print(f"      ERROR: {e}")
        
        ws.column_dimensions['A'].width = 80
    
    doc.close()
    wb.save(output_path)
    print(f"\nSaved Excel to: {output_path}")
    print(f"File size: {os.path.getsize(output_path)} bytes")

if __name__ == "__main__":
    # Create a test PDF with text and an image
    test_pdf = "debug_excel_test.pdf"
    test_xlsx = "debug_excel_test.xlsx"
    
    # Create dummy PDF with image
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((50, 50), "Test Document", fontsize=20)
    page.insert_text((50, 100), "This is a test line", fontsize=12)
    
    # Add a simple colored rectangle as "image" content
    rect = fitz.Rect(50, 150, 250, 300)
    page.draw_rect(rect, color=(0, 0, 1), fill=(0.9, 0.9, 1))
    page.insert_text((60, 200), "Blue Box", fontsize=14)
    
    doc.save(test_pdf)
    doc.close()
    print(f"Created test PDF: {test_pdf}")
    
    # Now test conversion
    test_pdf_to_excel_with_images(test_pdf, test_xlsx)
