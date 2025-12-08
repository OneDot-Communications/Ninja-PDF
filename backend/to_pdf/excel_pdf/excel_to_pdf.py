import os
import openpyxl
import xlrd
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch


def _workbook_to_table_data_openpyxl(wb, max_rows_per_sheet=500):
    """Extract data from openpyxl workbook as list of (sheet_name, table_data) tuples"""
    sheets_data = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        table_data = []
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if row_count >= max_rows_per_sheet:
                break
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            # Convert all cells to strings, replace None with empty string
            clean_row = [str(cell) if cell is not None else "" for cell in row]
            table_data.append(clean_row)
            row_count += 1
        if table_data:
            sheets_data.append((sheet_name, table_data))
    return sheets_data


def _workbook_to_table_data_xlrd(wb, max_rows_per_sheet=500):
    """Extract data from xlrd workbook as list of (sheet_name, table_data) tuples"""
    sheets_data = []
    for sheet in wb.sheets():
        table_data = []
        rows_to_process = min(sheet.nrows, max_rows_per_sheet)
        for row_idx in range(rows_to_process):
            row = sheet.row(row_idx)
            clean_row = [str(cell.value) if cell.value is not None else "" for cell in row]
            if not any(val.strip() for val in clean_row):
                continue
            table_data.append(clean_row)
        if table_data:
            sheets_data.append((sheet.name, table_data))
    return sheets_data


def convert_excel_to_pdf(input_path, output_path):
    """
    Convert Excel (.xlsx or .xls) to PDF without LibreOffice/subprocess.
    - Tries openpyxl first (handles real .xlsx and many mislabeled .xls files)
    - Falls back to xlrd for true legacy .xls files
    - Generates simple HTML and renders to PDF with xhtml2pdf
    """

    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Input file not found: {input_path}")

    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    file_ext = os.path.splitext(input_path)[1].lower()
    sheets_data = None

    # Try openpyxl first (covers .xlsx and many mislabeled .xls files)
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
        sheets_data = _workbook_to_table_data_openpyxl(wb, max_rows_per_sheet=500)
    except Exception as openpyxl_error:
        # If extension is .xls, try legacy reader
        if file_ext == '.xls':
            try:
                wb = xlrd.open_workbook(input_path)
                sheets_data = _workbook_to_table_data_xlrd(wb, max_rows_per_sheet=500)
            except Exception as xlrd_error:
                raise RuntimeError(
                    "Failed to read Excel file. If this is a legacy .xls, "
                    "please resave as .xlsx and try again. "
                    f"Details: openpyxl: {openpyxl_error}; xlrd: {xlrd_error}"
                )
        else:
            raise RuntimeError(f"Failed to read Excel file: {openpyxl_error}")

    if not sheets_data:
        raise RuntimeError("No data found in the Excel file.")

    # Build PDF using reportlab directly
    try:
        doc = SimpleDocTemplate(output_path, pagesize=landscape(letter))
        story = []
        styles = getSampleStyleSheet()
        
        for sheet_idx, (sheet_name, table_data) in enumerate(sheets_data):
            if sheet_idx > 0:
                story.append(PageBreak())
            
            # Add sheet title
            story.append(Paragraph(f"<b>{sheet_name}</b>", styles['Heading2']))
            story.append(Spacer(1, 0.2 * inch))
            
            if not table_data:
                story.append(Paragraph("(Empty sheet)", styles['Normal']))
                continue
            
            # Limit columns to prevent overflow
            max_cols = 20
            truncated_data = [row[:max_cols] for row in table_data]
            
            # Create table with fixed column widths
            num_cols = max(len(row) for row in truncated_data) if truncated_data else 1
            col_width = min(1.2 * inch, (doc.width / num_cols) if num_cols > 0 else 1 * inch)
            
            table = Table(truncated_data, colWidths=[col_width] * num_cols, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(table)
            story.append(Spacer(1, 0.3 * inch))
        
        doc.build(story)
    except Exception as e:
        raise RuntimeError(f"Excel to PDF conversion failed: {str(e)}")

    if not os.path.exists(output_path) or os.path.getsize(output_path) == 0:
        raise RuntimeError("PDF file was not created or is empty")

    return output_path
